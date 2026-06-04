from openpyxl import Workbook

from test_chest_cartilage_month_value_conservation import load_calculation_namespace


def _fill_rgb(cell):
    fill = cell.fill
    if fill is None or fill.fill_type != "solid":
        return None
    color = fill.fgColor
    return color.rgb or color.indexed or color.theme


def test_apply_row_min_max_highlight_marks_numeric_extremes_per_row():
    ns = load_calculation_namespace()
    wb = Workbook()
    ws = wb.active

    ws.append(["含税单价", "5月18日", "5月19日", "5月20日", "5月21日"])
    ws.append(["腿类", 2.0, 5.0, 5.0, 3.0])
    ws.append(["胸类", 4.0, None, 1.0, ""])
    ws.append(["翅类", 7.0, 7.0, 7.0, 7.0])

    ns["_apply_row_min_max_highlight"](
        ws=ws,
        col_start=2,
        col_end=5,
        data_start_row=2,
        data_end_row=4,
    )

    green = ns["TREND_MAX_FILL_COLOR"]
    yellow = ns["TREND_MIN_FILL_COLOR"]

    assert _fill_rgb(ws.cell(2, 2)) == yellow
    assert _fill_rgb(ws.cell(2, 3)) == green
    assert _fill_rgb(ws.cell(2, 4)) == green
    assert _fill_rgb(ws.cell(2, 5)) is None

    assert _fill_rgb(ws.cell(3, 2)) == green
    assert _fill_rgb(ws.cell(3, 4)) == yellow
    assert _fill_rgb(ws.cell(3, 3)) is None
    assert _fill_rgb(ws.cell(3, 5)) is None

    for col_idx in range(2, 6):
        assert _fill_rgb(ws.cell(4, col_idx)) is None
